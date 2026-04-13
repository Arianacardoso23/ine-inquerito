import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import date

st.set_page_config(page_title="Ponto de Situação do Inquérito", page_icon="📊", layout="wide")
 
# ============================================
# AUTENTICAÇÃO SIMPLES (SEM BANCO DE DADOS)
# ============================================
# Podes adicionar mais utilizadores aqui
USERS = {
   "mariatavares": "admin123",
    "helgabarros": "admin123",
    "joseborges": "admin123",
    "elgatavares": "admin123",
    "alicepinto": "admin123",
    "adilsonvarela": "admin123",
    "aryanacardoso": "admin123",
    "janecasfortes": "admin123",
    "arianatavares": "admin123",
    "aliciamota": "admin123",
    "teresamoniz": "admin123"
}

def check_password():
    """Retorna True se o utilizador estiver autenticado."""
    
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    
    if st.session_state.authenticated:
        return True
    
    # Formulário de login
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("""
            <div style="text-align: center; padding: 2rem; background-color: #f0f2f6; border-radius: 10px; margin-top: 3rem;">
                <h1 style="color: #1a3c6e;">📊 Ponto de Situação do Inquérito</h1>
                <h3 style="color: #666;">Sistema de Monitorização — INE Cabo Verde</h3>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        with st.form("login_form"):
            st.markdown("### 🔐 Iniciar Sessão")
            username = st.text_input("Utilizador", placeholder="Digite o seu username")
            password = st.text_input("Palavra-passe", type="password", placeholder="Digite a sua password")
            
            col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
            with col_btn2:
                submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")
            
            if submitted:
                if username in USERS and USERS[username] == password:
                    st.session_state.authenticated = True
                    st.session_state.username = username
                    st.rerun()
                else:
                    st.error("❌ Utilizador ou palavra-passe incorretos!")
        
        st.markdown("""
            <div style="text-align: center; color: #666; font-size: 0.8rem; margin-top: 2rem;">
                <p>Se não tiver credenciais, contacte o administrador do sistema.</p>
            </div>
        """, unsafe_allow_html=True)
    
    return False

# Verificar autenticação antes de mostrar conteúdo
if not check_password():
    st.stop()

# ============================================
# CONTEÚDO PRINCIPAL (APÓS LOGIN)
# ============================================

# Mensagem de boas-vindas na sidebar
with st.sidebar:
    st.markdown(f"""
        ### 👤 Bem-vindo(a)""")
    st.divider()


ILHAS = {1:"Santo Antão",2:"São Vicente",3:"São Nicolau",4:"Sal",5:"Boavista",6:"Maio",7:"Santiago",8:"Fogo",9:"Brava"}
CONCELHOS = {11:"Ribeira Grande",12:"Paul",13:"Porto Novo",21:"São Vicente",31:"Ribeira Brava",32:"Tarrafal de São Nicolau",41:"Sal",51:"Boavista",61:"Maio",71:"Tarrafal",72:"Santa Catarina",73:"Santa Cruz",74:"Praia",75:"S. Domingos",76:"S. Miguel",77:"S. Salvador do Mundo",78:"S. Lourenço dos Órgãos",79:"Ribeira Grande de Santiago",81:"Mosteiro",82:"São Filipe",83:"Santa Catarina do Fogo",91:"Brava"}
 
FOLHA_CONCELHO = {
    'RG':11,'PL':12,'PN':13,'SV':21,'RB':31,'TSN':32,'SL':41,'BV':51,
    'MA':61,'TA':71,'SC':72,'SZ':73,'PR':74,'SD':75,'SM':76,'SSM':77,
    'SLO':78,'RGST':79,'MOST':81,'SF':82,'SCF':83,'BR':91
}
 
st.markdown('<h1 style="color:#1a3c6e;border-bottom:3px solid #1a3c6e">📊 Ponto de Situação do Inquérito — INE Cabo Verde</h1>', unsafe_allow_html=True)
 
# ── Upload ficheiros ───────────────────────────────────────────────────────────
col_u1, col_u2 = st.columns(2)
with col_u1:
    uploaded = st.file_uploader("📂 Ficheiro de extração (Excel .xlsx)", type=["xlsx"])
with col_u2:
    uploaded_amostra = st.file_uploader("🎯 Ficheiro de amostra prevista (Excel .xlsx)", type=["xlsx"])
 
if uploaded is None:
    st.info("👆 Carregue o ficheiro de extração para visualizar o ponto de situação.")
    st.stop()
 
@st.cache_data
def load_amostra(amostra_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(amostra_bytes), read_only=True)
    resumo = []
    for folha, cod in FOLHA_CONCELHO.items():
        if folha in wb.sheetnames:
            ws = wb[folha]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = sum(1 for r in rows if any(v is not None for v in r))
            nome = CONCELHOS.get(cod, folha)
            resumo.append({"cod_concelho": cod, "concelho_nome": nome, "amostra_prevista": total})
    return pd.DataFrame(resumo)
 
@st.cache_data
def load_data(file_bytes):
    df     = pd.read_excel(BytesIO(file_bytes), sheet_name="Alojamento")
    df_ind = pd.read_excel(BytesIO(file_bytes), sheet_name="Individuo")
 
    df["agente"]     = df["USER_CREATE"].astype(str).replace("None","Desconhecido").str.split("@").str[0].str.replace("."," ",regex=False).str.title()
    df_ind["agente"] = df_ind["USER_CREATE"].astype(str).replace("None","Desconhecido").str.split("@").str[0].str.replace("."," ",regex=False).str.title()
 
    df["ilha_nome"]     = df["cod_ilha"].map(ILHAS).fillna(df["cod_ilha"].astype(str))
    df["concelho_nome"] = df["cod_concelho"].map(CONCELHOS).fillna(df["cod_concelho"].astype(str))
 
    df["ponto_valido"]     = (df["AA0200"]==1).astype(int)
    df["ponto_invalido"]   = (df["AA0200"]!=1).astype(int)
    df["res_habitual"]     = (df["AA0302"]==1).astype(int)
    df["secundaria"]       = (df["AA0302"]==2).astype(int)
    df["vazio"]            = (df["AA0302"]==3).astype(int)
    df["outros_fins"]      = (df["AA0302"]==4).astype(int)
    df["inacessivel"]      = (df["AA0302"]==5).astype(int)
    df["outra_situacao"]   = (df["AA0302"]==6).astype(int)
    df["recusa"]           = (df["AA0401"]==7).astype(int)
    df["agreg_inquiridos"] = (df["AA0605"]==1).astype(int)
    df["data_aloj"]        = pd.to_datetime(df["dt_creation_module_3"], errors="coerce").dt.date
    df_ind["data_ind"]     = pd.to_datetime(df_ind["dt_creation_module_232"], errors="coerce").dt.date
 
    ind_por_agente = df_ind.groupby("agente").agg(individuos_inquiridos=("REFERENCIA","count")).reset_index()
    df = df.merge(ind_por_agente[["agente","individuos_inquiridos"]], on="agente", how="left")
    df["individuos_inquiridos"] = df["individuos_inquiridos"].fillna(0).astype(int)
 
    return df, df_ind
 
file_bytes_main = uploaded.read()
df, df_ind = load_data(file_bytes_main)
 
# Carregar amostra
df_amostra = None
if uploaded_amostra:
    df_amostra = load_amostra(uploaded_amostra.read())
 
# # ── Filtros ────────────────────────────────────────────────────────────────────
col_f1, col_f2, col_f3 = st.columns(3)
 
with col_f1:
    ilha_sel = st.selectbox("🏝️ Selecionar Ilha", ["Todas"] + sorted(df["ilha_nome"].unique()))
 
df_filt = df if ilha_sel=="Todas" else df[df["ilha_nome"]==ilha_sel]
 
with col_f2:
    concelho_sel = st.selectbox("🏙️ Selecionar Concelho", ["Todos"] + sorted(df_filt["concelho_nome"].unique()))
 
if concelho_sel != "Todos":
    df_filt = df_filt[df_filt["concelho_nome"]==concelho_sel]
 
with col_f3:
    agente_sel = st.selectbox("👤 Selecionar Inquiridor", ["Todos"] + sorted(df_filt["agente"].unique()))
 
if agente_sel != "Todos":
    df_filt = df_filt[df_filt["agente"]==agente_sel]
    df_ind_filt = df_ind[df_ind["agente"]==agente_sel]
else:
    df_ind_filt = df_ind[df_ind["agente"].isin(df_filt["agente"].unique())]
 
st.divider()
 
# ── KPIs ───────────────────────────────────────────────────────────────────────
total_aloj    = len(df_filt)
total_validos = int(df_filt["ponto_valido"].sum())
total_inv     = int(df_filt["ponto_invalido"].sum())
total_agentes = df_filt["agente"].nunique()
total_ind     = len(df_ind_filt)
 
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("🏠 Total Alojamentos",    f"{total_aloj:,}")
k2.metric("✅ Pontos Válidos",        f"{total_validos:,}")
k3.metric("❌ Pontos Inválidos",      f"{total_inv:,}")
k4.metric("👤 Agentes Ativos",        f"{total_agentes}")
k5.metric("👥 Indivíduos Inquiridos", f"{total_ind:,}")
 
st.divider()
 
# ── Quadro resumo por agente ───────────────────────────────────────────────────
st.subheader("📋 Quadro de Pontos feitos pelos Agentes de Terreno")
 
resumo = df_filt.groupby(["ilha_nome","concelho_nome","agente"], dropna=False).agg(
    Total=("REFERENCIA","count"),
    Pontos_Validos=("ponto_valido","sum"),
    Pontos_Invalidos=("ponto_invalido","sum"),
    Res_Habitual=("res_habitual","sum"),
    Secundaria_Sazonal=("secundaria","sum"),
    Vazio=("vazio","sum"),
    Outros_Fins=("outros_fins","sum"),
    Inacessivel=("inacessivel","sum"),
    Outra_Situacao=("outra_situacao","sum"),
    Recusa=("recusa","sum"),
    Agreg_Inquiridos=("agreg_inquiridos","sum"),
    Individuos_Inquiridos=("individuos_inquiridos","max"), 
).reset_index()
 
linha_total = pd.DataFrame([{
    "ilha_nome":"TOTAL","concelho_nome":"","agente":"",
    "Total":resumo["Total"].sum(),"Pontos_Validos":resumo["Pontos_Validos"].sum(),
    "Pontos_Invalidos":resumo["Pontos_Invalidos"].sum(),"Res_Habitual":resumo["Res_Habitual"].sum(),
    "Secundaria_Sazonal":resumo["Secundaria_Sazonal"].sum(),"Vazio":resumo["Vazio"].sum(),
    "Outros_Fins":resumo["Outros_Fins"].sum(),"Inacessivel":resumo["Inacessivel"].sum(),
    "Outra_Situacao":resumo["Outra_Situacao"].sum(),"Recusa":resumo["Recusa"].sum(),
    "Agreg_Inquiridos":resumo["Agreg_Inquiridos"].sum(),
    "Individuos_Inquiridos":resumo["Individuos_Inquiridos"].sum(),
}])
resumo_display = pd.concat([resumo, linha_total], ignore_index=True)
resumo_display.columns = ["Ilha","Concelho","Agente","Total","Pontos Válidos","Pontos Inválidos",
    "Residência Habitual","Secundária/Sazonal","Vazio","Ocupados Outros Fins",
    "Alojamento Inacessível","Outra Situação","Recusa","Agregados Inquiridos","Indivíduos Inquiridos"]
 
def highlight_total(row):
    if row["Ilha"]=="TOTAL":
        return ["background-color:#1a3c6e;color:white;font-weight:bold"]*len(row)
    return [""]*len(row)
 
st.dataframe(resumo_display.style.apply(highlight_total,axis=1), use_container_width=True, hide_index=True)
 
# ── Evolução diária por inquiridor ─────────────────────────────────────────────
if agente_sel != "Todos":
    st.divider()
 
    st.markdown("**📆 Detalhe diário**")
    det = df_filt.groupby("data_aloj").agg(
        Total=("REFERENCIA","count"),Validos=("ponto_valido","sum"),Invalidos=("ponto_invalido","sum"),
        Res_Habitual=("res_habitual","sum"),Secundaria=("secundaria","sum"),Vazio=("vazio","sum"),
        Outros_Fins=("outros_fins","sum"),Inacessivel=("inacessivel","sum"),
        Outra_Situacao=("outra_situacao","sum"),Recusa=("recusa","sum"),Agregados=("agreg_inquiridos","sum"),
    ).reset_index().sort_values("data_aloj",ascending=False)
    det.columns=["Data","Total","Válidos","Inválidos","Res. Habitual","Secundária","Vazio","Outros Fins","Inacessível","Outra Situação","Recusa","Agregados"]
    st.dataframe(det, use_container_width=True, hide_index=True)
 
st.divider()
 
# ── Gráficos gerais ────────────────────────────────────────────────────────────
st.subheader("📊 Análise Visual")
tab1, tab2 = st.tabs(["Progresso por Agente","Distribuição de Alojamentos"])
with tab1:
    rs = resumo.sort_values("Pontos_Validos",ascending=True)
    fig = go.Figure()
    fig.add_trace(go.Bar(x=rs["Pontos_Validos"],y=rs["agente"],orientation="h",name="Válidos",marker_color="#1a3c6e",text=rs["Pontos_Validos"],textposition="auto"))
    fig.add_trace(go.Bar(x=rs["Pontos_Invalidos"],y=rs["agente"],orientation="h",name="Inválidos",marker_color="#e57373",text=rs["Pontos_Invalidos"],textposition="auto"))
    fig.update_layout(barmode="stack",title="Pontos por Agente",xaxis_title="Nº de Alojamentos",yaxis_title="Agente",height=400)
    st.plotly_chart(fig, use_container_width=True)
    
with tab2:
    labels = ["Res. Habitual", "Secundária/Sazonal", "Vazio", "Outros Fins", "Inacessível", "Outra Situação", "Recusa"]
    values = [int(df_filt[c].sum()) for c in ["res_habitual", "secundaria", "vazio", "outros_fins", "inacessivel", "outra_situacao", "recusa"]]
    
    fig2 = go.Figure(go.Pie(
        values=values,
        labels=labels,
        marker=dict(colors=["#1a3c6e", "#a8dadc", "#2e86ab", "#e57373", "#f4a261", "#e76f51", "#e9c46a"]),
        hole=0.4,
        textinfo="percent",  # Apenas percentagem dentro das fatias
        textposition="inside",  # Posicionar dentro das fatias
        textfont=dict(color="white", size=12),  # Texto branco para contraste
        hoverinfo="label+percent+value",  # Informação ao passar o mouse
        showlegend=True  # Mostrar legenda ao lado
    ))
    
    fig2.update_layout(
        title="Distribuição por Tipo de Alojamento",
        height=400,
        legend=dict(
            orientation="v",  # Vertical
            yanchor="middle",
            y=0.5,
            xanchor="left",
            x=1.05  # Posicionar à direita do gráfico
        ),
        margin=dict(l=50, r=150, t=50, b=50)  # Margem direita maior para a legenda
    )
    
    st.plotly_chart(fig2, use_container_width=True)
 
st.divider()
 
# ── Exportar ───────────────────────────────────────────────────────────────────
st.subheader("⬇️ Exportar Dados")
 
# ── Download da amostra com estado (feito / por fazer) ─────────────────────────
if uploaded_amostra:
    st.markdown("##### 📋 Amostra com Estado dos Pontos")
 
    @st.cache_data
    def gerar_amostra_estado(amostra_bytes, extracao_bytes):
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(amostra_bytes), read_only=True)
 
        # Todos os IDs já visitados na extração
        df_ext = pd.read_excel(BytesIO(extracao_bytes), sheet_name="Alojamento")
        df_ext["agente"] = df_ext["USER_CREATE"].astype(str).replace("None","Desconhecido").str.split("@").str[0].str.replace("."," ",regex=False).str.title()
        ids_feitos = set(df_ext["REFERENCIA"].astype(str).str.strip())
        ref_agente = dict(zip(df_ext["REFERENCIA"].astype(str).str.strip(), df_ext["agente"]))
        ref_data   = dict(zip(df_ext["REFERENCIA"].astype(str).str.strip(), df_ext["dt_creation_module_3"].astype(str)))
 
        output = BytesIO()
        todas_folhas = []
 
        for folha, cod in FOLHA_CONCELHO.items():
            if folha not in wb.sheetnames:
                continue
            ws = wb[folha]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            df_f = pd.DataFrame(rows, columns=["ID_EDIFICIO","ID_ALOJAMENTO","DR","CIDADE_ZONA","BAIRRO_LUGAR","OBSERVACAO"])
            df_f = df_f.dropna(subset=["ID_ALOJAMENTO"])
            df_f["ID_ALOJAMENTO"] = df_f["ID_ALOJAMENTO"].astype(str).str.strip()
            df_f["CONCELHO"]      = CONCELHOS.get(cod, folha)
            df_f["ESTADO"]        = df_f["ID_ALOJAMENTO"].apply(lambda x: "Feito" if x in ids_feitos else "Por Fazer")
            df_f["AGENTE"]        = df_f["ID_ALOJAMENTO"].map(ref_agente).fillna("")
            df_f["DATA_RECOLHA"]  = df_f["ID_ALOJAMENTO"].map(ref_data).fillna("")
            todas_folhas.append(df_f)
 
        df_total = pd.concat(todas_folhas, ignore_index=True)
        df_total = df_total[["CONCELHO","ID_EDIFICIO","ID_ALOJAMENTO","DR","CIDADE_ZONA","BAIRRO_LUGAR","ESTADO","AGENTE","DATA_RECOLHA","OBSERVACAO"]]
        df_total.columns = ["Concelho","ID Edifício","ID Alojamento","DR","Cidade/Zona","Bairro/Lugar","Estado","Agente","Data Recolha","Observação"]
 
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Folha resumo
            resumo_estado = df_total.groupby("Concelho").agg(
                Total=("ID Alojamento","count"),
                Feitos=("Estado", lambda x: (x=="Feito").sum()),
                Por_Fazer=("Estado", lambda x: (x=="Por Fazer").sum()),
            ).reset_index()
            resumo_estado.columns = ["Concelho","Total Amostra","Feitos","Por Fazer"]
            # Linha total
            lt = pd.DataFrame([{
                "Concelho": "TOTAL",
                "Total Amostra": resumo_estado["Total Amostra"].sum(),
                "Feitos": resumo_estado["Feitos"].sum(),
                "Por Fazer": resumo_estado["Por Fazer"].sum(),
                
            }])
            resumo_estado = pd.concat([resumo_estado, lt], ignore_index=True)
            resumo_estado.to_excel(writer, index=False, sheet_name="Resumo")
 
            # Folha com todos os pontos
            df_total.to_excel(writer, index=False, sheet_name="Todos os Pontos")
 
            # Folha só com os pontos por fazer
            por_fazer = df_total[df_total["Estado"]==" Por Fazer"]
            por_fazer.to_excel(writer, index=False, sheet_name="Por Fazer")
 
            # Folha só com os pontos feitos
            feitos = df_total[df_total["Estado"]==" Feito"]
            feitos.to_excel(writer, index=False, sheet_name="Feitos")
 
        return output.getvalue()
 
    # Filtrar por concelho/ilha se selecionado
col_am1, col_am2, col_am3 = st.columns(3)

with col_am1:
    st.metric("📦 Total Amostra",
        f"{int(df_amostra['amostra_prevista'].sum()):,}" if df_amostra is not None else "—")

with col_am2:
    if df_amostra is not None:
        total_feitos_geral = int(df["ponto_valido"].sum()) + int(df["ponto_invalido"].sum())
        st.metric("✅ Feitos", f"{total_feitos_geral:,}")

with col_am3:
    if df_amostra is not None:
        total_prev = int(df_amostra["amostra_prevista"].sum())
        total_feit = int(df["ponto_valido"].sum()) + int(df["ponto_invalido"].sum())
        st.metric("⏳ Por Fazer", f"{max(total_prev - total_feit, 0):,}")

# Botão em baixo, fora das colunas
excel_amostra = gerar_amostra_estado(uploaded_amostra.getvalue(), file_bytes_main)
st.download_button(
    "📥 Descarregar Amostra com Estado (Feito / Por Fazer)",
    data=excel_amostra,
    file_name="amostra_estado_pontos.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
    
)
st.caption("O ficheiro contém 4 folhas: Resumo · Todos os Pontos · Por Fazer · Feitos")
 
st.divider()
 
col_d1, col_d2 = st.columns(2)
 
@st.cache_data
def to_excel_df(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resumo")
    return output.getvalue()
 
nome = "ponto_situacao_geral" if ilha_sel=="Todas" and concelho_sel=="Todos" else f"ponto_situacao_{(concelho_sel if concelho_sel!='Todos' else ilha_sel).replace(' ','_')}"
 
with col_d1:
    st.download_button("📊 Exportar Quadro por Agente",data=to_excel_df(resumo_display),file_name=f"{nome}_agentes.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
 
with col_d2:
    rt = df_filt.groupby(["ilha_nome","concelho_nome"]).agg(
        Total=("REFERENCIA","count"),Pontos_Validos=("ponto_valido","sum"),Pontos_Invalidos=("ponto_invalido","sum"),
        Res_Habitual=("res_habitual","sum"),Secundaria_Sazonal=("secundaria","sum"),Vazio=("vazio","sum"),
        Outros_Fins=("outros_fins","sum"),Inacessivel=("inacessivel","sum"),Outra_Situacao=("outra_situacao","sum"),
        Recusa=("recusa","sum"),Agregados=("agreg_inquiridos","sum"),Individuos=("individuos_inquiridos","max"),
    ).reset_index()
    if df_amostra is not None:
        rt = rt.merge(df_amostra[["concelho_nome","amostra_prevista"]], on="concelho_nome", how="left")
        rt["pct_meta"] = (rt["Pontos_Validos"] / rt["amostra_prevista"] * 100).round(1)
    lt2 = {c: rt[c].sum() if rt[c].dtype in ["int64","float64"] else ("TOTAL" if c=="ilha_nome" else "") for c in rt.columns}
    rt = pd.concat([rt, pd.DataFrame([lt2])], ignore_index=True)
    st.download_button("📋 Exportar Resumo por Ilha/Concelho",data=to_excel_df(rt),file_name=f"{nome}_territorio.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
 
st.divider()
# ============================================
# BOTÃO SAIR NO FINAL
# ============================================
st.markdown("---")
col_rodape1, col_rodape2, col_rodape3 = st.columns([1, 2, 1])

with col_rodape2:
    if st.button("🚪 Terminar Sessão", use_container_width=True, type="secondary"):
        # Limpar estado de autenticação
        st.session_state.authenticated = False
        st.session_state.username = None
        st.rerun()
 
st.caption("Sistema de Monitorização do Inquérito — INE Cabo Verde")
